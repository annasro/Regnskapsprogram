from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4
from openpyxl import load_workbook
#funksjon som lager regningen
def create_invoice(*parameters):
    print(parameters)
    #import excel file
    wb = load_workbook("faktura_mal.xlsx", data_only = True)
    #åpne riktig ark
    sheet = wb['faktura']

    #lage pdf
    c = canvas.Canvas('./faktura/'+'faktura_' + str(fakturanummer)+"_"+ str(fakturadato) + '.pdf')
    #sidestørrelse
    c.setPageSize(A4)

    i = 2
    #verdier fra xcel ark
    #fakturadato_    = sheet.cell(row = i, column = 6).value
    #forfallsdato_   = sheet.cell(row = i, column = 7).value
    #leveringsdato_  = sheet.cell(row = i, column = 8).value
    kommentar        = sheet.cell(row = i, column = 9).value
    no_elever        = sheet.cell(row = 4, column = 2).value


    #logo
    #c.drawInlineImage('logo.jpg', start_x, start_y, width = bilde_bredde, height = bilde_høyde)

    #sidestørrelse A4
    width, height = A4
    margin = 40
    pdfmetrics.registerFont(TTFont('Arial','Arial.ttf'))

    #elementposisjoner
    v =  margin
    v1 = v +  margin*3
    v2 = v1 + margin*2
    v3 = v2 + margin*2
    v4 = v3 + margin*2
    orgno_bredde =stringWidth(str(b_organisasjonsnummer),'Helvetica',10)
    h = width - 3.5*margin - orgno_bredde
    h1 = width - 2*margin
    midt = width/2.0
    y =  height - margin

    #bedrift
    c.setFont('Helvetica-Bold', 10)
    c.drawString(v,y, str(b_navn))
    c.setFont('Helvetica', 10)

    c.drawString(h ,y,'Org.No.'+ str(b_organisasjonsnummer))
    y -= margin*0.4
    c.drawString(v,y, str(b_adresse))
    y -= margin*0.4
    c.drawString(v,y, str(b_postnummer))
    y -= margin*0.4
    #Faktura
    c.setFont('Arial', 15)
    c.drawString(h, y, 'Faktura')
    y -= margin

    c.setFont('Helvetica', 10)
    c.drawString(v,y, str(b_epost))
    c.drawString(h,y,'Kundenr.: ')
    c.drawString(h1,y, str(k_nummer))
    y -= margin*.45
    c.drawString(v,y, 'Kundetelefon: '+ str(b_telefon))
    c.drawString(h,y,'Fakturanr.: ')
    c.drawString(h1,y, str(fakturanummer))
    y -= margin*.4


    #fakurainfo
    c.drawString(h,y,'Fakturadato: ')
    c.drawString(h1,y, str(fakturadato))
    y -= margin*.4
    c.drawString(h,y,'Forfallsdato: ')
    c.setFont('Helvetica-Bold', 10)
    c.drawString(h1,y, str(forfallsdato))
    y -= margin*.4
    c.setFont('Helvetica', 10)
    c.drawString(h,y,'Leveringsdato: ')
    c.drawString(h1,y, str(leveringsdato))
    y -= margin

    #kunde
    c.setFont('Helvetica-Bold', 10)
    c.drawString(v,y, str(k_navn))
    y -= margin*.4
    c.setFont('Helvetica', 10)
    c.drawString(v,y, str(k_adresse))
    y -= margin*.4
    c.drawString(v,y, str(k_postnummer))
    y -= margin*2
    y = height -350


    #midtdel
    c.drawString(v,y,   'Beskrivelse')
    c.drawString(v1,y,  'Antall')
    c.drawString(v2,y,  'Timepris' )
    c.drawString(v3,y,  'Bonus' )
    c.drawString(h,y,   'MVA sats:' )
    c.drawString(h1,y,  'Netto pris')
    y -= margin*.4

    stop = 6 + no_elever
    for j in range (6, stop):
        c.setFont('Helvetica', 10)
        beskrivelse     = sheet.cell(row = j, column = 1).value
        timeantall      = sheet.cell(row = j, column = 2).value
        timepris        = sheet.cell(row = j, column = 3).value
        bonus           = sheet.cell(row = j, column = 4).value
        nettopris       = sheet.cell(row = j, column = 5).value
        moms_pros       = sheet.cell(row = j, column = 12).value
        sum             = sheet.cell(row = j, column = 14).value

        c.drawString(v,y, str(beskrivelse))
        c.drawString(v1,y,str(timeantall))
        c.drawString(v2,y,str(timepris))
        c.drawString(v3,y,str(bonus))

        c.setFont('Times-Roman', 10)
        c.drawString(h,y,  '0.00 %')
        c.drawString(h1,y, str(nettopris))
        y -= margin*.4

    pris_moms  = sheet.cell(row = stop, column = 7).value
    totalt     = sheet.cell(row = stop, column = 8).value

    y -= margin *2
    c.setFont('Helvetica-Bold', 10)
    c.drawString(h,y,'MVA:')
    c.setFont('Times-Roman', 10)
    c.drawString(h1,y, str(pris_moms))
    y -= margin*.4
    c.setFont('Helvetica-Bold', 10)
    c.drawString(h,y,'Total:')
    c.setFont('Times-Roman', 10)
    c.drawString(h1,y,str(totalt))
    y -= margin

    c.setFont('Helvetica', 10)
    c.drawString(v,y, 'Kommentar: ' + str(kommentar))


    #nederste del av faktura
    y = height - 650
    c.setFont('Times-Roman', 10)
    c.drawString(midt,y, str(totalt))
    y -= margin*1
    c.setFont('Helvetica-Bold', 10)
    c.drawString(h1,y, str(forfallsdato))
    c.setFont('Helvetica', 10)
    c.drawString(v,y,'Kundenr.: ')
    c.drawString(v1,y, str(k_nummer))
    y -= margin*0.4
    c.drawString(v,y,'Fakturanr.: ')
    c.drawString(v1,y, str(fakturanummer))
    y -= margin

    c.drawString(v,y, str(b_navn))
    c.drawString(h,y, str(k_navn))
    y -= margin*.4
    c.drawString(v,y, str(b_adresse))
    c.drawString(h,y, str(k_adresse))
    y -= margin*.4
    c.drawString(v,y, str(b_postnummer))
    c.drawString(h,y, str(k_postnummer))
    y -= margin

    c.setFont('Times-Roman', 10)
    c.drawString(midt,y,str(totalt))
    c.setFont('Helvetica', 10)
    c.drawString(h,y,'Kontomummer: ' + str(b_kontonummer))

    c.save()
