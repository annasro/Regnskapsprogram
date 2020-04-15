import sys
from openpyxl import load_workbook

def generate_from_document(document):
    # Loading worksheet
    wb = load_workbook(filename = document, data_only = True)
    sheet = wb.active

    # Allocating dictionaries to hold data for students and expenses
    students = {}
    expenses = {}

    # Looping over every row in sheet (excluding the header)
    for student, hours, expense, category in sheet.iter_rows(min_row = 2, max_row = sheet.max_row, values_only = True):
        # Checks if row contains student data
        if student:
            # If student key doesn't exist in dict, make it
            if student not in students.keys():
                students[student] = 0
            students[student] += hours # Add the hours in the row to the total for this student
        
        # Checks if row contains expense data
        elif expense:
            # If category key doesn't exist in dict, make it
            if category not in expenses.keys():
                expenses[category] = 0
            expenses[category] += expense # Add the expense in the row to the total for this category.

    print(students)
    print(expenses)

if __name__ == "__main__":
    generate_from_document(sys.argv[1])